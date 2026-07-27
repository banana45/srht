from __future__ import annotations

import csv
import html
import re
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import BinaryIO, Callable


@dataclass(frozen=True)
class ContractRow:
    party_a: str
    start_date: str
    end_date: str
    signing_date: str = ""


@dataclass(frozen=True)
class StoreProofRow:
    enterprise_name: str
    business_license: str
    account_name: str
    shop_url: str
    proof_date: str


@dataclass(frozen=True)
class AuthorizationLetterRow:
    enterprise_name: str
    authorization_date: str


HEADER_ALIASES = {
    "party_a": {"party_a", "甲方", "甲方名称", "甲方名字", "客户名称"},
    "start_date": {"start_date", "开始日期", "合同开始日期", "起始日期"},
    "end_date": {"end_date", "结束日期", "合同结束日期", "截止日期"},
    "signing_date": {"signing_date", "签署日期", "签订日期", "日期"},
}

STORE_PROOF_HEADER_ALIASES = {
    "enterprise_name": {"enterprise_name", "企业名", "企业名称", "公司名称", "认证主体"},
    "business_license": {"business_license", "营业执照", "营业执照号", "统一社会信用代码"},
    "account_name": {"account_name", "账户名称", "账号名称", "店铺名称"},
    "shop_url": {"shop_url", "店铺地址", "店铺链接", "店铺URL", "店铺url"},
    "proof_date": {"proof_date", "时间", "日期", "证明时间", "证明日期"},
}

STORE_PROOF_TEMPLATE_VALUES = {
    "enterprise_name": "厦门淑莱汝信息科技有限公司",
    "business_license": "91350206MAKGH6NDXJ",
    "account_name": "厦门淑莱汝网络设计",
    "shop_url": "https://xmslr.qidian.hzshengruikj.cn/",
    "proof_date": "2026年7月19日",
}

AUTHORIZATION_LETTER_HEADER_ALIASES = {
    "enterprise_name": {"enterprise_name", "企业名", "企业名称", "公司名称", "被授权人"},
    "authorization_date": {"authorization_date", "时间", "日期", "授权时间", "授权日期"},
}


def format_chinese_date(raw: str) -> str:
    value = str(raw or "").strip()
    match = re.fullmatch(r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})日?", value)
    if not match:
        raise ValueError("日期格式应为 2026-7-1、2026/7/1 或 2026年7月1日")

    year, month, day = (int(part) for part in match.groups())
    try:
        parsed = date(year, month, day)
    except ValueError as exc:
        raise ValueError("日期格式无效，请检查年月日") from exc
    return f"{parsed.year}年{parsed.month}月{parsed.day}日"


def normalize_row(raw: dict[str, object]) -> ContractRow:
    party_a = str(raw.get("party_a", "") or "").strip()
    start_date = str(raw.get("start_date", "") or "").strip()
    end_date = str(raw.get("end_date", "") or "").strip()
    signing_date = str(raw.get("signing_date", "") or "").strip()

    if not party_a:
        raise ValueError("甲方名称不能为空")
    if not start_date:
        raise ValueError("开始日期不能为空")
    if not end_date:
        raise ValueError("结束日期不能为空")

    return ContractRow(
        party_a=party_a,
        start_date=format_chinese_date(start_date),
        end_date=format_chinese_date(end_date),
        signing_date=format_chinese_date(signing_date) if signing_date else "",
    )


def normalize_store_proof_row(raw: dict[str, object]) -> StoreProofRow:
    enterprise_name = str(raw.get("enterprise_name", "") or "").strip()
    business_license = str(raw.get("business_license", "") or "").strip()
    account_name = str(raw.get("account_name", "") or "").strip()
    shop_url = str(raw.get("shop_url", "") or "").strip()
    proof_date = str(raw.get("proof_date", "") or "").strip()

    if not enterprise_name:
        raise ValueError("企业名不能为空")
    if not business_license:
        raise ValueError("营业执照不能为空")
    if not account_name:
        raise ValueError("账户名称不能为空")
    if not shop_url:
        raise ValueError("店铺地址不能为空")
    if not proof_date:
        raise ValueError("时间不能为空")

    return StoreProofRow(
        enterprise_name=enterprise_name,
        business_license=business_license,
        account_name=account_name,
        shop_url=shop_url,
        proof_date=format_chinese_date(proof_date),
    )


def normalize_authorization_letter_row(raw: dict[str, object]) -> AuthorizationLetterRow:
    enterprise_name = str(raw.get("enterprise_name", "") or "").strip()
    authorization_date = str(raw.get("authorization_date", "") or "").strip()

    if not enterprise_name:
        raise ValueError("企业名不能为空")
    if not authorization_date:
        raise ValueError("时间不能为空")

    return AuthorizationLetterRow(
        enterprise_name=enterprise_name,
        authorization_date=format_chinese_date(authorization_date),
    )


def parse_csv_rows(file_obj: BinaryIO) -> list[dict[str, str]]:
    content = file_obj.read()
    text = content.decode("utf-8-sig") if isinstance(content, bytes) else content

    reader = csv.DictReader(text.splitlines())
    if not reader.fieldnames:
        return []

    mapped_headers = {_canonical_header(header): header for header in reader.fieldnames}
    rows: list[dict[str, str]] = []
    for source in reader:
        rows.append(
            {
                key: str(source.get(mapped_headers.get(key, ""), "") or "").strip()
                for key in ("party_a", "start_date", "end_date", "signing_date")
            }
        )
    return rows


def parse_csv_store_proof_rows(file_obj: BinaryIO) -> list[dict[str, str]]:
    content = file_obj.read()
    text = content.decode("utf-8-sig") if isinstance(content, bytes) else content
    return _parse_csv_text(text, STORE_PROOF_HEADER_ALIASES, _store_proof_keys())


def parse_csv_authorization_letter_rows(file_obj: BinaryIO) -> list[dict[str, str]]:
    content = file_obj.read()
    text = content.decode("utf-8-sig") if isinstance(content, bytes) else content
    return _parse_csv_text(text, AUTHORIZATION_LETTER_HEADER_ALIASES, _authorization_letter_keys())


def parse_xlsx_rows(file_obj: BinaryIO) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []

    mapped_indexes = {_canonical_header(str(header or "")): index for index, header in enumerate(headers)}
    rows: list[dict[str, str]] = []
    for source in rows_iter:
        if not any(source):
            continue
        rows.append(
            {
                key: _cell_value(source[mapped_indexes[key]])
                if key in mapped_indexes and mapped_indexes[key] < len(source)
                else ""
                for key in ("party_a", "start_date", "end_date", "signing_date")
            }
        )
    workbook.close()
    return rows


def parse_xlsx_store_proof_rows(file_obj: BinaryIO) -> list[dict[str, str]]:
    return _parse_xlsx_rows(file_obj, STORE_PROOF_HEADER_ALIASES, _store_proof_keys())


def parse_xlsx_authorization_letter_rows(file_obj: BinaryIO) -> list[dict[str, str]]:
    return _parse_xlsx_rows(file_obj, AUTHORIZATION_LETTER_HEADER_ALIASES, _authorization_letter_keys())


def generate_contract(
    template_path: Path,
    output_path: Path,
    row: ContractRow,
    progress: Callable[[int, str], None] | None = None,
) -> None:
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, 5, "读取模板")

    with zipfile.ZipFile(template_path, "r") as source, zipfile.ZipFile(
        output_path, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "word/document.xml":
                _emit(progress, 35, "填充合同内容")
                xml = data.decode("utf-8")
                data = _fill_document_xml(xml, row).encode("utf-8")
            target.writestr(item, data)

    _emit(progress, 100, "完成")


def generate_contract_archive(
    template_path: Path,
    output_path: Path,
    rows: list[ContractRow],
    progress: Callable[[int, int, str], None] | None = None,
) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        generated_paths: list[Path] = []
        total = len(rows)
        for index, row in enumerate(rows, start=1):
            safe_name = _safe_filename(row.party_a) or f"contract-{index}"
            target = tmp_path / f"{index:03d}-{safe_name}.docx"

            def on_contract_progress(percent: int, message: str, row_index: int = index) -> None:
                if progress:
                    progress(row_index, percent, message)

            generate_contract(template_path, target, row, progress=on_contract_progress)
            generated_paths.append(target)
            if progress:
                progress(index, 100, f"已完成 {index}/{total}")

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as archive:
            for path in generated_paths:
                archive.write(path, path.name)


def generate_store_proof(
    template_path: Path,
    output_path: Path,
    row: StoreProofRow,
    progress: Callable[[int, str], None] | None = None,
) -> None:
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, 5, "读取模板")

    replacements = {
        STORE_PROOF_TEMPLATE_VALUES["enterprise_name"]: row.enterprise_name,
        STORE_PROOF_TEMPLATE_VALUES["business_license"]: row.business_license,
        STORE_PROOF_TEMPLATE_VALUES["account_name"]: row.account_name,
        STORE_PROOF_TEMPLATE_VALUES["shop_url"]: row.shop_url,
        STORE_PROOF_TEMPLATE_VALUES["proof_date"]: row.proof_date,
    }

    with zipfile.ZipFile(template_path, "r") as source, zipfile.ZipFile(
        output_path, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "word/document.xml":
                _emit(progress, 35, "填充经营证明内容")
                xml = data.decode("utf-8")
                data = _replace_document_text(xml, replacements).encode("utf-8")
            target.writestr(item, data)

    _emit(progress, 100, "完成")


def generate_store_proof_archive(
    template_path: Path,
    output_path: Path,
    rows: list[StoreProofRow],
    progress: Callable[[int, int, str], None] | None = None,
) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        generated_paths: list[Path] = []
        total = len(rows)
        for index, row in enumerate(rows, start=1):
            safe_name = _safe_filename(row.enterprise_name) or f"store-proof-{index}"
            target = tmp_path / f"{index:03d}-{safe_name}-店铺经营证明.docx"

            def on_proof_progress(percent: int, message: str, row_index: int = index) -> None:
                if progress:
                    progress(row_index, percent, message)

            generate_store_proof(template_path, target, row, progress=on_proof_progress)
            generated_paths.append(target)
            if progress:
                progress(index, 100, f"已完成 {index}/{total}")

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as archive:
            for path in generated_paths:
                archive.write(path, path.name)


def generate_authorization_letter(
    template_path: Path,
    output_path: Path,
    row: AuthorizationLetterRow,
    progress: Callable[[int, str], None] | None = None,
) -> None:
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, 5, "读取模板")

    with zipfile.ZipFile(template_path, "r") as source, zipfile.ZipFile(
        output_path, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "word/document.xml":
                _emit(progress, 35, "填充授权函内容")
                xml = data.decode("utf-8")
                data = _fill_authorization_letter_xml(xml, row).encode("utf-8")
            target.writestr(item, data)

    _emit(progress, 100, "完成")


def generate_authorization_letter_archive(
    template_path: Path,
    output_path: Path,
    rows: list[AuthorizationLetterRow],
    progress: Callable[[int, int, str], None] | None = None,
) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        generated_paths: list[Path] = []
        total = len(rows)
        for index, row in enumerate(rows, start=1):
            safe_name = _safe_filename(row.enterprise_name) or f"authorization-letter-{index}"
            target = tmp_path / f"{index:03d}-{safe_name}-授权函.docx"

            def on_letter_progress(percent: int, message: str, row_index: int = index) -> None:
                if progress:
                    progress(row_index, percent, message)

            generate_authorization_letter(template_path, target, row, progress=on_letter_progress)
            generated_paths.append(target)
            if progress:
                progress(index, 100, f"已完成 {index}/{total}")

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as archive:
            for path in generated_paths:
                archive.write(path, path.name)


def _fill_document_xml(xml: str, row: ContractRow) -> str:
    def replace_paragraph(match: re.Match[str]) -> str:
        paragraph = match.group(0)
        text = _paragraph_text(paragraph)

        if "甲方：" in text and "_______________" in text:
            return _replace_after_label(paragraph, "甲方：", row.party_a)

        if "合同履行期限自" in text and "至" in text:
            return _replace_contract_term(paragraph, row.start_date, row.end_date)

        if row.signing_date and "日期：" in text and "_______________" in text:
            return _replace_after_label(paragraph, "日期：", row.signing_date)

        return paragraph

    return re.sub(r"<w:p[\s\S]*?</w:p>", replace_paragraph, xml)


def _fill_authorization_letter_xml(xml: str, row: AuthorizationLetterRow) -> str:
    fill_next_authorizer_date = False

    def replace_paragraph(match: re.Match[str]) -> str:
        nonlocal fill_next_authorizer_date
        paragraph = match.group(0)
        text = _paragraph_text(paragraph)

        if "现授权人依法授权" in text and "（下称" in text:
            paragraph = _replace_between_labels(paragraph, "现授权人依法授权", "（下称", row.enterprise_name)
            return _strip_underlines(paragraph)

        if text.strip().startswith("授权人（盖章）"):
            fill_next_authorizer_date = True
            return paragraph

        if fill_next_authorizer_date and "时间：" in text:
            fill_next_authorizer_date = False
            return _replace_after_label_whitespace(paragraph, "时间：", row.authorization_date)

        return paragraph

    return re.sub(r"<w:p[\s\S]*?</w:p>", replace_paragraph, xml)


def _replace_after_label(paragraph: str, label: str, replacement: str) -> str:
    text = _paragraph_text(paragraph)
    start = text.find(label)
    if start == -1:
        return paragraph

    blank_start = text.find("_", start + len(label))
    if blank_start == -1:
        return paragraph

    blank_end = blank_start
    while blank_end < len(text) and text[blank_end] == "_":
        blank_end += 1
    paragraph = _replace_text_range(paragraph, blank_start, blank_end, replacement, underline=False)
    return _strip_underlines(paragraph)


def _replace_after_label_whitespace(paragraph: str, label: str, replacement: str) -> str:
    text = _paragraph_text(paragraph)
    start = text.find(label)
    if start == -1:
        return paragraph

    value_start = start + len(label)
    value_end = value_start
    while value_end < len(text) and text[value_end].isspace():
        value_end += 1
    return _replace_text_range(paragraph, value_start, value_end, replacement, underline=False)


def _replace_between_labels(paragraph: str, start_label: str, end_label: str, replacement: str) -> str:
    text = _paragraph_text(paragraph)
    start_label_pos = text.find(start_label)
    if start_label_pos == -1:
        return paragraph
    start = start_label_pos + len(start_label)
    end = text.find(end_label, start)
    if end == -1:
        return paragraph
    return _replace_text_range(paragraph, start, end, replacement, underline=False)


def _replace_contract_term(paragraph: str, start_date: str, end_date: str) -> str:
    text = _paragraph_text(paragraph)
    start_label_pos = text.find("自")
    middle_label_pos = text.find("至", start_label_pos + 1)
    if start_label_pos == -1 or middle_label_pos == -1:
        return paragraph

    first_start, first_end = _find_underline_range(text, start_label_pos, middle_label_pos)
    second_start, second_end = _find_underline_range(text, middle_label_pos, len(text))
    if second_start >= 0:
        paragraph = _replace_text_range(paragraph, second_start, second_end, end_date, underline=False)
    if first_start >= 0:
        paragraph = _replace_text_range(paragraph, first_start, first_end, start_date, underline=False)
    return _strip_underlines(paragraph)


def _find_underline_range(text: str, start: int, end: int) -> tuple[int, int]:
    segment = text[start:end]
    match = re.search(r"_+", segment)
    if not match:
        return -1, -1
    return start + match.start(), start + match.end()


def _replace_text_range(paragraph: str, start: int, end: int, replacement: str, underline: bool) -> str:
    result: list[str] = []
    xml_cursor = 0
    text_cursor = 0
    inserted = False

    for match in re.finditer(r"<w:r(?:\s[^>]*)?>[\s\S]*?</w:r>", paragraph):
        result.append(paragraph[xml_cursor : match.start()])
        run = match.group(0)
        run_value = _run_text(run)
        run_start = text_cursor
        run_end = text_cursor + len(run_value)
        text_cursor = run_end
        xml_cursor = match.end()

        if run_end <= start or run_start >= end:
            result.append(run)
            continue

        local_start = max(start - run_start, 0)
        local_end = min(end - run_start, len(run_value))
        prefix = run_value[:local_start]
        suffix = run_value[local_end:]

        if not inserted:
            if prefix:
                result.append(_set_run_text(run, prefix))
            replacement_run = _set_run_text(run, replacement)
            result.append(_ensure_run_underline(replacement_run) if underline else replacement_run)
            if suffix:
                result.append(_set_run_text(run, suffix))
            inserted = True
        elif suffix:
            result.append(_set_run_text(run, suffix))

    result.append(paragraph[xml_cursor:])
    return "".join(result)


def _paragraph_text(paragraph: str) -> str:
    return "".join(_xml_text_unescape(match.group(1)) for match in _text_matches(paragraph))


def _run_text(run: str) -> str:
    return "".join(_xml_text_unescape(match.group(1)) for match in _text_matches(run))


def _text_matches(xml: str):
    return re.finditer(r"<w:t(?:\s[^>]*)?>([\s\S]*?)</w:t>", xml)


def _set_run_text(run: str, value: str) -> str:
    escaped = html.escape(value, quote=False)
    attrs = ' xml:space="preserve"' if value[:1].isspace() or value[-1:].isspace() else ""
    text_tag = f"<w:t{attrs}>{escaped}</w:t>"
    first = re.search(r"<w:t(?:\s[^>]*)?>[\s\S]*?</w:t>", run)
    if not first:
        return run
    replaced = run[: first.start()] + text_tag + run[first.end() :]
    return replaced[: first.start() + len(text_tag)] + re.sub(
        r"<w:t(?:\s[^>]*)?>[\s\S]*?</w:t>",
        "",
        replaced[first.start() + len(text_tag) :],
    )


def _ensure_run_underline(run: str) -> str:
    if "<w:u " in run or "<w:u/>" in run:
        return run
    if "<w:rPr>" in run:
        if "<w:lang" in run:
            return re.sub(r"(<w:rPr>[\s\S]*?)(<w:lang\b)", r'\1<w:u w:val="single"/>\2', run, count=1)
        return run.replace("</w:rPr>", '<w:u w:val="single"/></w:rPr>', 1)
    return re.sub(r"(<w:r(?:\s[^>]*)?>)", r'\1<w:rPr><w:u w:val="single"/></w:rPr>', run, count=1)


def _strip_underlines(xml: str) -> str:
    return re.sub(r"<w:u(?:\s[^>]*)?/>", "", xml)


def _xml_text_unescape(value: str) -> str:
    return html.unescape(value)


def _replace_document_text(xml: str, replacements: dict[str, str]) -> str:
    def replace_paragraph(match: re.Match[str]) -> str:
        paragraph = match.group(0)
        for old, new in replacements.items():
            paragraph = _replace_visible_text(paragraph, old, new)
        return paragraph

    return re.sub(r"<w:p[\s\S]*?</w:p>", replace_paragraph, xml)


def _replace_visible_text(paragraph: str, old: str, new: str) -> str:
    text = _paragraph_text(paragraph)
    start = text.find(old)
    if start == -1:
        return paragraph
    return _replace_text_range(paragraph, start, start + len(old), new, underline=False)


def _parse_csv_text(
    text: str,
    aliases: dict[str, set[str]],
    keys: tuple[str, ...],
) -> list[dict[str, str]]:
    reader = csv.DictReader(text.splitlines())
    if not reader.fieldnames:
        return []

    mapped_headers = {_canonical_header(header, aliases): header for header in reader.fieldnames}
    rows: list[dict[str, str]] = []
    for source in reader:
        rows.append(
            {
                key: str(source.get(mapped_headers.get(key, ""), "") or "").strip()
                for key in keys
            }
        )
    return rows


def _parse_xlsx_rows(
    file_obj: BinaryIO,
    aliases: dict[str, set[str]],
    keys: tuple[str, ...],
) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []

    mapped_indexes = {_canonical_header(str(header or ""), aliases): index for index, header in enumerate(headers)}
    rows: list[dict[str, str]] = []
    for source in rows_iter:
        if not any(source):
            continue
        rows.append(
            {
                key: _cell_value(source[mapped_indexes[key]])
                if key in mapped_indexes and mapped_indexes[key] < len(source)
                else ""
                for key in keys
            }
        )
    workbook.close()
    return rows


def _store_proof_keys() -> tuple[str, ...]:
    return ("enterprise_name", "business_license", "account_name", "shop_url", "proof_date")


def _authorization_letter_keys() -> tuple[str, ...]:
    return ("enterprise_name", "authorization_date")


def _canonical_header(header: str | None, aliases: dict[str, set[str]] = HEADER_ALIASES) -> str:
    normalized = str(header or "").strip()
    for canonical, choices in aliases.items():
        if normalized in choices:
            return canonical
    return normalized


def _cell_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_filename(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", value).strip()


def _emit(callback: Callable[[int, str], None] | None, percent: int, message: str) -> None:
    if callback:
        callback(percent, message)
